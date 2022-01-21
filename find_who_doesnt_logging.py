from openpyxl import load_workbook

import change_to_at


def read_exist_list(file) -> dict[str, bool]:
    m: dict[str, bool] = {}
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        xm = row[0].value
        is_login = row[1].value is not None
        m[xm] = is_login
    return m


def read_total_list(file, d: dict[str, bool]) -> set[str]:
    s: set[str] = set()
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=2):
        xm = row[0].value
        if d.get(xm) is not None:
            if not d[xm]:
                s.add(xm)
    return s


if __name__ == '__main__':
    di = read_exist_list('./fxdj.xlsx')
    change_to_at.change_to_at('\n'.join(e for e in read_total_list('./14-1.20.xlsx', di)))
