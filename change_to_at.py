import re
from openpyxl import load_workbook


def change_to_at(s: str):
    s = s.strip()
    ss = s.split("\n")
    for i in range(len(ss)):
        ss[i] = "@" + ss[i]
    res = " ".join(ss)
    print(res + " ")


def read_ocr_file(file: str, compare_dict: dict[str, str]):
    finished_dict: dict[str, str] = {}
    with open(file, "r", encoding='utf-8') as f:
        for line in f.readlines():
            line = line.strip()
            re1 = r"^.{2,3}\s*[0-9]{10}$"
            re2 = r"^[0-9]{10}\s*.{2,3}$"
            if not (re.match(re1, line) or re.match(re2, line)):
                continue
            res = re.search(r"[0-9]{10}", line).group()
            print(res, compare_dict.get(res))
            finished_dict[res] = compare_dict.get(res)

    print("total：", len(compare_dict))
    print()
    print("finished:")
    print(finished_dict, "total:" + str(len(finished_dict)), sep="\n")
    print()
    print("unfinished:")
    unfinished_set = finished_dict.items() ^ compare_dict.items()
    print(unfinished_set, "total:" + str(len(unfinished_set)), sep="\n")
    print()
    unfinished_name = ""
    for elem in unfinished_set:
        unfinished_name += "@" + elem[1] + " "
    print(unfinished_name)


def read_excel(file) -> dict[str, str]:
    m: dict[str, str] = {}
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        xh = str(row[0].value)
        xm = row[1].value
        m[xh] = xm
    return m


if __name__ == "__main__":
    read_ocr_file("./QQ202112251919_OCR/result.txt",read_excel("./bjtxl.xlsx"))
    # read_ocr_file("./QQ202112111603_OCR/result.txt", read_excel("./bjtxl.xlsx"))